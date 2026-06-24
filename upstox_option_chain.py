"""
=============================================================================
 Upstox Live Option Chain Fetcher  |  Appends to a SINGLE open .xlsx
=============================================================================
 How saving works:
   - File is NOT open in Excel  →  openpyxl saves directly (fast)
   - File IS  open in Excel     →  xlwings writes new rows via COM
                                    (no file close / no new file created)

 Excel sheet layout:
   "Calls" sheet   – Fetch Time | Spot | Expiry | Strike | PCR | Call market data + Greeks
   "Puts"  sheet   – Same prefix, Put market data + Greeks
   "Greeks Guide"  – Plain-English explanation of every Greek

 Setup:
   pip install requests openpyxl xlwings

 Usage:
   1.  Set ACCESS_TOKEN  (or $env:UPSTOX_ACCESS_TOKEN)
   2.  python upstox_option_chain.py
=============================================================================
"""

import os
import sys
import time
import logging
import datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
#  USER CONFIGURATION
# =============================================================================

ACCESS_TOKEN: str = os.getenv("UPSTOX_ACCESS_TOKEN", "YOUR_ACCESS_TOKEN_HERE")

# NSE_INDEX|Nifty 50 | NSE_INDEX|Nifty Bank | NSE_INDEX|Nifty Fin Service
UNDERLYING: str = "NSE_INDEX|Nifty 50"

# "YYYY-MM-DD"  or  None  to auto-pick nearest expiry
EXPIRY_DATE = None

# Seconds between live refreshes
REFRESH_INTERVAL: int = 5

# Output file  — always ONE file, never creates extra copies
OUTPUT_FILE: str = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "option_chain_live.xlsx"
)

# Number of strikes around ATM to capture (None = all)
STRIKES_AROUND_ATM = 20

# =============================================================================
#  CONSTANTS
# =============================================================================

BASE_URL = "https://api.upstox.com/v2"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Colour palette (hex strings)
C_HDR_BG      = "1E1E2E"   # column-header row background
C_HDR_FG      = "CDD6F4"   # column-header row text
C_META_BG     = "EAF2FF"   # Fetch Time / Spot / Expiry / Strike / PCR
C_ATM_BG      = "FFF9C4"   # ATM row
C_ALT_BG      = "F8F9FA"   # OTM zebra (even rows)
C_ITM_CALL_BG = "A9DFBF"   # ITM call (strike < spot)
C_ITM_PUT_BG  = "F1948A"   # ITM put  (strike > spot)
C_OTM_CALL_BG = "D4EFDF"   # OTM call odd rows
C_OTM_PUT_BG  = "FADBD8"   # OTM put  odd rows
C_CALLS_TITLE = "1A5276"   # Calls sheet title bar
C_PUTS_TITLE  = "641E16"   # Puts  sheet title bar

# =============================================================================
#  COLUMN DEFINITIONS
# =============================================================================

_PREFIX = [
    ("Fetch Time", "_fetch_time",  19, "YYYY-MM-DD HH:MM:SS"),
    ("Spot",       "_spot",        10, "#,##0.00"),
    ("Expiry",     "_expiry",      12, "@"),
    ("Strike",     "strike_price", 11, "#,##0.00"),
    ("PCR",        "pcr",           8, "0.00"),
]
N_PREFIX = len(_PREFIX)

_CALL_DATA = [
    ("Volume",  "call_options.market_data.volume",     10, "#,##0"),
    ("OI",      "call_options.market_data.oi",         10, "#,##0"),
    ("Prev OI", "call_options.market_data.prev_oi",    10, "#,##0"),
    ("Chg OI",  "_call_chg_oi",                       10, "#,##0"),
    ("Bid Qty", "call_options.market_data.bid_qty",     9, "#,##0"),
    ("Bid",     "call_options.market_data.bid_price",   9, "#,##0.00"),
    ("Ask",     "call_options.market_data.ask_price",   9, "#,##0.00"),
    ("Ask Qty", "call_options.market_data.ask_qty",     9, "#,##0"),
    ("LTP",     "call_options.market_data.ltp",         9, "#,##0.00"),
    ("IV",      "call_options.option_greeks.iv",        8, "0.00"),
    ("Delta",   "call_options.option_greeks.delta",     8, "0.0000"),
    ("Gamma",   "call_options.option_greeks.gamma",     9, "0.00000"),
    ("Theta",   "call_options.option_greeks.theta",     9, "0.00"),
    ("Vega",    "call_options.option_greeks.vega",      8, "0.0000"),
    ("PoP %",   "call_options.option_greeks.pop",       8, "0.00"),
]

_PUT_DATA = [
    ("Volume",  "put_options.market_data.volume",      10, "#,##0"),
    ("OI",      "put_options.market_data.oi",          10, "#,##0"),
    ("Prev OI", "put_options.market_data.prev_oi",     10, "#,##0"),
    ("Chg OI",  "_put_chg_oi",                        10, "#,##0"),
    ("Bid Qty", "put_options.market_data.bid_qty",      9, "#,##0"),
    ("Bid",     "put_options.market_data.bid_price",    9, "#,##0.00"),
    ("Ask",     "put_options.market_data.ask_price",    9, "#,##0.00"),
    ("Ask Qty", "put_options.market_data.ask_qty",      9, "#,##0"),
    ("LTP",     "put_options.market_data.ltp",          9, "#,##0.00"),
    ("IV",      "put_options.option_greeks.iv",         8, "0.00"),
    ("Delta",   "put_options.option_greeks.delta",      8, "0.0000"),
    ("Gamma",   "put_options.option_greeks.gamma",      9, "0.00000"),
    ("Theta",   "put_options.option_greeks.theta",      9, "0.00"),
    ("Vega",    "put_options.option_greeks.vega",       8, "0.0000"),
    ("PoP %",   "put_options.option_greeks.pop",        8, "0.00"),
]

CALLS_COLS = _PREFIX + _CALL_DATA
PUTS_COLS  = _PREFIX + _PUT_DATA

# =============================================================================
#  API LAYER
# =============================================================================

def _headers():
    return {"Authorization": f"Bearer {ACCESS_TOKEN}", "Accept": "application/json"}


def fetch_expiries(instrument_key):
    url = f"{BASE_URL}/option/contract"
    try:
        r = requests.get(url, headers=_headers(),
                         params={"instrument_key": instrument_key}, timeout=15)
        r.raise_for_status()
        expiries = sorted(set(c["expiry"] for c in r.json().get("data", []) if c.get("expiry")))
        log.info("Found %d expiry dates for %s", len(expiries), instrument_key)
        return expiries
    except Exception as e:
        log.error("fetch_expiries error: %s", e)
        return []


def fetch_option_chain(instrument_key, expiry_date):
    url = f"{BASE_URL}/option/chain"
    r = requests.get(url, headers=_headers(),
                     params={"instrument_key": instrument_key, "expiry_date": expiry_date},
                     timeout=15)
    r.raise_for_status()
    body = r.json()
    if body.get("status") != "success":
        raise ValueError(f"API non-success: {body}")
    return body.get("data", [])


def get_nearest_expiry(instrument_key):
    expiries = fetch_expiries(instrument_key)
    today    = datetime.date.today().isoformat()
    upcoming = [e for e in expiries if e >= today]
    if not upcoming:
        raise RuntimeError("No upcoming expiries found.")
    chosen = upcoming[0]
    log.info("Auto-selected expiry: %s  (available: %s)", chosen, expiries[:5])
    return chosen

# =============================================================================
#  DATA HELPERS
# =============================================================================

def deep_get(obj, path):
    for part in path.split("."):
        if not isinstance(obj, dict):
            return None
        obj = obj.get(part)
    return obj


def enrich_row(row):
    row["_call_chg_oi"] = (deep_get(row, "call_options.market_data.oi")      or 0) - \
                          (deep_get(row, "call_options.market_data.prev_oi") or 0)
    row["_put_chg_oi"]  = (deep_get(row, "put_options.market_data.oi")       or 0) - \
                          (deep_get(row, "put_options.market_data.prev_oi")  or 0)
    return row


def filter_strikes(data, spot, n):
    if n is None or spot is None or not data:
        return data
    strikes = [r["strike_price"] for r in data]
    atm     = min(range(len(strikes)), key=lambda i: abs(strikes[i] - spot))
    return data[max(0, atm - n) : min(len(data) - 1, atm + n) + 1]


def get_spot_price(data):
    return data[0].get("underlying_spot_price") if data else None


def _atm_strike(data, spot):
    if spot and data:
        strikes = [r["strike_price"] for r in data]
        return min(strikes, key=lambda s: abs(s - spot))
    return None

# =============================================================================
#  COLOUR HELPERS
# =============================================================================

def _hex_to_rgb(h):
    """Convert '#RRGGBB' or 'RRGGBB' to (R, G, B) tuple."""
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

# =============================================================================
#  PATH A — openpyxl  (file not open in Excel)
# =============================================================================

def _op_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _op_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _op_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _op_write_header(ws, cols, title, title_bg):
    """Write two header rows via openpyxl. Called once per sheet."""
    thin  = _op_border()
    total = len(cols)

    # Row 1 — merged title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, color=C_HDR_FG, size=12, name="Calibri")
    c.fill      = _op_fill(title_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — column headers
    for ci, (label, _, width, _) in enumerate(cols, start=1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font      = _op_font(bold=True, color=C_HDR_FG)
        c.fill      = _op_fill(C_HDR_BG)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"


def _op_append_rows(ws, data, spot, expiry, fetched_at, cols, itm_below, itm_bg, otm_bg):
    """Append new rows via openpyxl."""
    thin  = _op_border()
    atm   = _atm_strike(data, spot)
    start = ws.max_row + 1

    for off, raw in enumerate(data):
        row               = enrich_row(raw)
        row["_fetch_time"] = fetched_at
        row["_spot"]       = spot
        row["_expiry"]     = expiry
        strike      = row.get("strike_price", 0)
        is_atm      = atm is not None and strike == atm
        is_itm      = spot is not None and (strike < spot if itm_below else strike > spot)
        row_idx     = start + off
        alt         = row_idx % 2 == 0

        for ci, (_, key, _, fmt) in enumerate(cols, start=1):
            c = ws.cell(row=row_idx, column=ci)
            c.value        = row.get(key) if key.startswith("_") else deep_get(row, key)
            c.number_format = fmt
            c.alignment    = Alignment(horizontal="center", vertical="center")
            c.border       = thin
            c.font         = _op_font(bold=is_atm, size=9)

            is_meta = ci <= N_PREFIX
            if is_atm:
                c.fill = _op_fill(C_ATM_BG)
            elif is_meta:
                c.fill = _op_fill(C_META_BG)
            elif is_itm:
                c.fill = _op_fill(itm_bg)
            elif alt:
                c.fill = _op_fill(C_ALT_BG)
            else:
                c.fill = _op_fill(otm_bg)

        ws.row_dimensions[row_idx].height = 15

    return len(data)


def _op_ensure_greeks(wb):
    if "Greeks Guide" in wb.sheetnames:
        return
    ws = wb.create_sheet("Greeks Guide")
    rows = [
        ("Greek",    "Meaning",                                                        "Typical Range"),
        ("Delta",    "Change in option price per ₹1 move in underlying",               "0→1 (call) | -1→0 (put)"),
        ("Gamma",    "Rate of change of Delta per ₹1 move",                           "Always +ve; peaks at ATM"),
        ("Theta",    "Daily time decay — option price lost per day",                   "Always -ve (₹/day)"),
        ("Vega",     "Price change per 1% change in Implied Volatility",               "Always +ve"),
        ("IV",       "Implied Volatility — market's future vol expectation",           "% annualised"),
        ("PoP %",    "Probability of Profit at expiry",                                "0–100 % (≈50 at ATM)"),
        ("PCR",      "Put-Call OI Ratio — >1 bearish, <1 bullish",                    ">1 / <1"),
        ("Chg OI",   "Change in Open Interest vs. previous close",                     "+ve = new longs added"),
        ("ITM Call", "Strike < Spot  →  call is in-the-money  (green)",               "—"),
        ("ITM Put",  "Strike > Spot  →  put  is in-the-money  (red)",                 "—"),
        ("ATM",      "Strike nearest to Spot  (yellow row)",                           "—"),
    ]
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    for ri, rd in enumerate(rows, start=1):
        for ci, val in enumerate(rd, start=1):
            cell            = ws.cell(row=ri, column=ci, value=val)
            cell.font       = _op_font(bold=(ri == 1),
                                       color=C_HDR_FG if ri == 1 else "000000", size=10)
            if ri == 1:
                cell.fill   = _op_fill(C_HDR_BG)
            cell.alignment  = Alignment(wrap_text=True, vertical="center")
            cell.border     = _op_border()
        ws.row_dimensions[ri].height = 20


def _load_or_create(filepath):
    """Load existing workbook or create a new one with Calls + Puts sheets."""
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            is_new = False
        except Exception as e:
            log.warning("Could not open existing file (%s) — creating fresh.", e)
            wb = openpyxl.Workbook()
            is_new = True
    else:
        wb = openpyxl.Workbook()
        is_new = True

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]

    calls_new = puts_new = False
    if "Calls" not in wb.sheetnames:
        wb.create_sheet("Calls", 0)
        calls_new = True
    if "Puts" not in wb.sheetnames:
        idx = (wb.sheetnames.index("Calls") + 1) if "Calls" in wb.sheetnames else 1
        wb.create_sheet("Puts", idx)
        puts_new = True

    return wb, wb["Calls"], wb["Puts"], (calls_new or is_new), (puts_new or is_new)


def _openpyxl_run(expiry, data, spot, fetched_at):
    """Full openpyxl path: load → append → save."""
    wb, calls_ws, puts_ws, cn, pn = _load_or_create(OUTPUT_FILE)

    title_c = f"CALLS  |  {UNDERLYING}  |  Expiry: {expiry}"
    title_p = f"PUTS   |  {UNDERLYING}  |  Expiry: {expiry}"

    if cn or calls_ws.max_row < 2:
        _op_write_header(calls_ws, CALLS_COLS, title_c, C_CALLS_TITLE)
    if pn or puts_ws.max_row < 2:
        _op_write_header(puts_ws, PUTS_COLS, title_p, C_PUTS_TITLE)

    nc = _op_append_rows(calls_ws, data, spot, expiry, fetched_at,
                         CALLS_COLS, True,  C_ITM_CALL_BG, C_OTM_CALL_BG)
    np_ = _op_append_rows(puts_ws,  data, spot, expiry, fetched_at,
                          PUTS_COLS,  False, C_ITM_PUT_BG,  C_OTM_PUT_BG)
    _op_ensure_greeks(wb)

    wb.save(OUTPUT_FILE)
    log.info("Saved  -> %s  (Calls: %d rows | Puts: %d rows)",
             OUTPUT_FILE, calls_ws.max_row, puts_ws.max_row)
    return True

# =============================================================================
#  PATH B — xlwings  (file is open in Excel — write via COM, NO file touching)
# =============================================================================

def _xw_last_row(xw_sheet):
    """Return the last used row index (1-based). Returns 0 for empty sheet."""
    try:
        used = xw_sheet.api.UsedRange
        return used.Row + used.Rows.Count - 1
    except Exception:
        return 0


def _xw_write_header(xw_sheet, cols, title, title_bg_hex):
    """Write two header rows directly via xlwings COM."""
    n = len(cols)
    xl_center = -4108  # xlCenter

    # Row 1 — merged title
    try:
        xw_sheet.range((1, 1), (1, n)).api.Merge()
    except Exception:
        pass
    xw_sheet.range(1, 1).value = title
    xw_sheet.range((1, 1), (1, n)).color     = _hex_to_rgb(title_bg_hex)
    xw_sheet.range(1, 1).font.bold           = True
    xw_sheet.range(1, 1).font.size           = 12
    xw_sheet.range(1, 1).font.color         = _hex_to_rgb(C_HDR_FG)
    xw_sheet.range(1, 1).api.HorizontalAlignment = xl_center
    xw_sheet.api.Rows(1).RowHeight = 22

    # Row 2 — column labels
    xw_sheet.range(2, 1).value = [[col[0] for col in cols]]
    hdr = xw_sheet.range((2, 1), (2, n))
    hdr.color                     = _hex_to_rgb(C_HDR_BG)
    hdr.font.bold                 = True
    hdr.font.color                = _hex_to_rgb(C_HDR_FG)
    hdr.api.HorizontalAlignment   = xl_center
    hdr.api.WrapText              = True
    xw_sheet.api.Rows(2).RowHeight = 28

    # Column widths
    for ci, (_, _, width, _) in enumerate(cols, start=1):
        xw_sheet.api.Columns(ci).ColumnWidth = width

    # Freeze at row 3
    try:
        xw_sheet.activate()
        win = xw_sheet.book.app.api.ActiveWindow
        win.FreezePanes = False
        xw_sheet.range("A3").api.Select()
        win.FreezePanes = True
    except Exception:
        pass


def _xw_ensure_headers(xw_sheet, cols, title, title_bg_hex):
    """Write header rows only if the sheet is empty or missing them."""
    try:
        if xw_sheet.range(2, 1).value == cols[0][0]:
            return   # headers already present
    except Exception:
        pass
    _xw_write_header(xw_sheet, cols, title, title_bg_hex)


def _xw_append_rows(xw_sheet, data, spot, expiry, fetched_at, cols, itm_below, itm_bg, otm_bg):
    """
    Append rows directly to an open Excel sheet via xlwings COM.
    Uses bulk value write (1 COM call) + range-based colour application.
    """
    xl_center = -4108   # xlCenter constant

    last_row  = _xw_last_row(xw_sheet)
    start_row = last_row + 1
    n_cols    = len(cols)
    atm       = _atm_strike(data, spot)

    # ── Enrich all rows and build value matrix ────────────────────────────────
    enriched = []
    matrix   = []
    for raw in data:
        row               = enrich_row(raw)
        row["_fetch_time"] = fetched_at
        row["_spot"]       = spot
        row["_expiry"]     = expiry
        enriched.append(row)
        matrix.append([
            row.get(key) if key.startswith("_") else deep_get(row, key)
            for (_, key, _, _) in cols
        ])

    # ── 1 COM call: write all values at once ──────────────────────────────────
    xw_sheet.range(start_row, 1).value = matrix

    # ── Number formats (one call per column) ──────────────────────────────────
    end_row = start_row + len(data) - 1
    for ci, (_, _, _, fmt) in enumerate(cols, start=1):
        xw_sheet.range(start_row, ci).resize(len(data), 1).number_format = fmt
    xw_sheet.range((start_row, 1), (end_row, n_cols)) \
            .api.HorizontalAlignment = xl_center

    # ── Borders for the entire new block (1 COM call) ─────────────────────────
    block = xw_sheet.range((start_row, 1), (end_row, n_cols)).api
    block.Borders.LineStyle = 1   # xlContinuous
    block.Borders.Weight    = 2   # xlThin

    # ── Row height (1 COM call) ───────────────────────────────────────────────
    block.RowHeight = 15

    # ── Categorise rows for colour grouping ──────────────────────────────────
    atm_rows = []
    itm_rows = []
    alt_rows = []
    otm_rows = []

    for off, row in enumerate(enriched):
        strike  = row.get("strike_price", 0)
        row_idx = start_row + off
        is_atm  = atm is not None and strike == atm
        is_itm  = spot is not None and (strike < spot if itm_below else strike > spot)
        alt     = row_idx % 2 == 0

        if is_atm:
            atm_rows.append(row_idx)
        elif is_itm:
            itm_rows.append(row_idx)
        elif alt:
            alt_rows.append(row_idx)
        else:
            otm_rows.append(row_idx)

    # ── Apply colours by group (one range call per row) ───────────────────────
    data_start = N_PREFIX + 1

    # Meta columns (same for all rows)
    for ri in range(start_row, start_row + len(data)):
        xw_sheet.range((ri, 1), (ri, N_PREFIX)).color = _hex_to_rgb(C_META_BG)

    # Data-area colours
    for ri in itm_rows:
        xw_sheet.range((ri, data_start), (ri, n_cols)).color = _hex_to_rgb(itm_bg)
    for ri in alt_rows:
        xw_sheet.range((ri, data_start), (ri, n_cols)).color = _hex_to_rgb(C_ALT_BG)
    for ri in otm_rows:
        xw_sheet.range((ri, data_start), (ri, n_cols)).color = _hex_to_rgb(otm_bg)

    # ATM — override entire row + bold
    for ri in atm_rows:
        xw_sheet.range((ri, 1), (ri, n_cols)).color     = _hex_to_rgb(C_ATM_BG)
        xw_sheet.range((ri, 1), (ri, n_cols)).font.bold = True

    return len(data)


def _xw_ensure_greeks(book):
    """Add Greeks Guide sheet directly via xlwings if missing."""
    if "Greeks Guide" in [s.name for s in book.sheets]:
        return
    try:
        ws = book.sheets.add("Greeks Guide")
        rows = [
            ("Greek",    "Meaning",                                                  "Typical Range"),
            ("Delta",    "Change in option price per ₹1 move in underlying",         "0→1 (call) | -1→0 (put)"),
            ("Gamma",    "Rate of change of Delta per ₹1 move",                     "Always +ve"),
            ("Theta",    "Daily time decay",                                         "Always -ve (₹/day)"),
            ("Vega",     "Price change per 1% IV change",                            "Always +ve"),
            ("IV",       "Implied Volatility",                                       "% annualised"),
            ("PoP %",    "Probability of Profit at expiry",                          "0–100 %"),
            ("PCR",      "Put-Call OI Ratio",                                        ">1 bearish | <1 bullish"),
            ("Chg OI",   "Change in OI vs. previous close",                          "+ve = new longs"),
            ("ITM Call", "Strike < Spot (green)",                                    "—"),
            ("ITM Put",  "Strike > Spot (red)",                                      "—"),
            ("ATM",      "Strike nearest Spot (yellow)",                             "—"),
        ]
        ws.range(1, 1).value = rows
        ws.range((1, 1), (1, 3)).color     = _hex_to_rgb(C_HDR_BG)
        ws.range((1, 1), (1, 3)).font.bold = True
        ws.range((1, 1), (1, 3)).font.color = _hex_to_rgb(C_HDR_FG)
        ws.api.Columns(1).ColumnWidth = 14
        ws.api.Columns(2).ColumnWidth = 60
        ws.api.Columns(3).ColumnWidth = 30
    except Exception as e:
        log.debug("Greeks Guide via xlwings skipped: %s", e)


def _xlwings_run(expiry, data, spot, fetched_at):
    """Write directly to the open Excel workbook via xlwings COM. No file I/O."""
    try:
        import xlwings as xw  # type: ignore
    except ImportError:
        log.error("xlwings not installed — run: pip install xlwings")
        return False

    # Find the open workbook — search all open books by name or full path
    filename = os.path.basename(OUTPUT_FILE)
    book = None
    try:
        for b in xw.books:
            if b.name.lower() == filename.lower() or \
               b.fullname.lower() == OUTPUT_FILE.lower():
                book = b
                break
    except Exception:
        pass

    if book is None:
        # File is locked but not found via COM — open it (xlwings will find existing instance)
        try:
            book = xw.Book(OUTPUT_FILE)
        except Exception as e:
            log.error("Cannot open via xlwings: %s", e)
            return False

    sheet_cfg = [
        ("Calls", CALLS_COLS, True,  C_ITM_CALL_BG, C_OTM_CALL_BG,
         f"CALLS  |  {UNDERLYING}  |  Expiry: {expiry}", C_CALLS_TITLE),
        ("Puts",  PUTS_COLS,  False, C_ITM_PUT_BG,  C_OTM_PUT_BG,
         f"PUTS   |  {UNDERLYING}  |  Expiry: {expiry}", C_PUTS_TITLE),
    ]

    for sname, cols, itm_below, itm_bg, otm_bg, title, title_bg in sheet_cfg:
        try:
            xws = book.sheets[sname]
        except Exception:
            xws = book.sheets.add(sname)

        _xw_ensure_headers(xws, cols, title, title_bg)
        n = _xw_append_rows(xws, data, spot, expiry, fetched_at,
                             cols, itm_below, itm_bg, otm_bg)
        log.info("xlwings: +%d rows → '%s'  (total row %d)", n, sname, _xw_last_row(xws))

    _xw_ensure_greeks(book)
    book.save()
    log.info("Saved via xlwings -> %s", OUTPUT_FILE)
    return True

# =============================================================================
#  FILE LOCK CHECK
# =============================================================================

def _is_locked(filepath):
    """True if the file exists and is currently locked (e.g., open in Excel)."""
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "r+b"):
            return False
    except (PermissionError, IOError):
        return True

# =============================================================================
#  MAIN LOOP
# =============================================================================

def run_once(expiry):
    fetched_at = datetime.datetime.now()
    log.info("Fetching | %s | expiry=%s", UNDERLYING, expiry)

    try:
        data = fetch_option_chain(UNDERLYING, expiry)
    except requests.HTTPError as e:
        log.error("HTTP error: %s", e)
        return False
    except Exception as e:
        log.error("Fetch error: %s", e)
        return False

    if not data:
        log.warning("Empty response — market may be closed.")
        return False

    spot = get_spot_price(data)
    log.info("Spot: %s  |  Strikes fetched: %d", spot, len(data))
    data = filter_strikes(data, spot, STRIKES_AROUND_ATM)

    if _is_locked(OUTPUT_FILE):
        # File is open in Excel → write directly via COM (no file I/O)
        return _xlwings_run(expiry, data, spot, fetched_at)
    else:
        # File is free → use openpyxl (full formatting, fast save)
        return _openpyxl_run(expiry, data, spot, fetched_at)


def main():
    if ACCESS_TOKEN == "YOUR_ACCESS_TOKEN_HERE":
        log.error(
            "\n  ACCESS_TOKEN not set!\n\n"
            "  PowerShell:\n"
            "    $env:UPSTOX_ACCESS_TOKEN = 'your_token'\n"
            "    python upstox_option_chain.py\n"
        )
        sys.exit(1)

    expiry = EXPIRY_DATE
    if not expiry:
        try:
            expiry = get_nearest_expiry(UNDERLYING)
        except Exception as e:
            log.error("Could not determine expiry: %s", e)
            sys.exit(1)

    log.info("=" * 65)
    log.info("  Upstox Live Option Chain")
    log.info("  Underlying : %s", UNDERLYING)
    log.info("  Expiry     : %s", expiry)
    log.info("  Refresh    : every %d s", REFRESH_INTERVAL)
    log.info("  Output     : %s", OUTPUT_FILE)
    log.info("  Strikes    : %s around ATM", STRIKES_AROUND_ATM or "ALL")
    log.info("=" * 65)

    run_once(expiry)
    while True:
        log.info("Next refresh in %d s …", REFRESH_INTERVAL)
        time.sleep(REFRESH_INTERVAL)
        run_once(expiry)


if __name__ == "__main__":
    main()
