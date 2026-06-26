"""
=============================================================================
 Upstox Live Option Chain Fetcher  |  Appends to a SINGLE open .xlsx
=============================================================================
 How saving works:
   - File is NOT open in Excel  →  openpyxl saves directly (fast)
   - File IS  open in Excel     →  xlwings writes new rows via COM
                                    (no file close / no new file created)

  Excel sheet layout:
    "Option Chain" sheet – Fetch Time | Expiry | Call data | Strike | Put data
    "Greeks Guide"       – Plain-English explanation of every Greek

 Setup:
   pip install requests openpyxl xlwings

 Usage:
   1.  Set ACCESS_TOKEN  (or $env:UPSTOX_ACCESS_TOKEN)
   2.  python upstox_option_chain.py
=============================================================================
"""

import os
import sys
import time
import logging
import datetime
import requests
import openpyxl
import email.utils
import webbrowser
import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

# Load .env file manually if it exists
def _load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip()

_load_env()

# =============================================================================
#  USER CONFIGURATION
# =============================================================================

def get_access_token() -> str:
    _load_env()
    return os.getenv("UPSTOX_ACCESS_TOKEN", "YOUR_ACCESS_TOKEN_HERE")

# Upstox API Credentials for OAuth Token Exchange
CLIENT_ID: str = os.getenv("UPSTOX_CLIENT_ID", "54a272c8-4978-432b-a599-ffc4b32b9e89")
CLIENT_SECRET: str = os.getenv("UPSTOX_CLIENT_SECRET", "ejvyac8mhy")
REDIRECT_URI: str = os.getenv("UPSTOX_REDIRECT_URI", "https://www.google.com/")

# Global state to track if current token is unauthorized
token_invalid: bool = False

# NSE_INDEX|Nifty 50 | NSE_INDEX|Nifty Bank | NSE_INDEX|Nifty Fin Service
UNDERLYING: str = "NSE_INDEX|Nifty 50"

# "YYYY-MM-DD"  or  None  to auto-pick nearest expiry
EXPIRY_DATE = None

# Seconds between live refreshes
REFRESH_INTERVAL: int = 5

# Output file  — always ONE file, never creates extra copies
OUTPUT_FILE: str = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "option_chain_live.xlsx"
)
HISTORY_FILE: str = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "option_chain_history.xlsx"
)

# Number of strikes around ATM to capture (None = all)
STRIKES_AROUND_ATM = 20

# =============================================================================
#  CONSTANTS
# =============================================================================

BASE_URL = "https://api.upstox.com/v2"

# Setup logging to both console and file
_log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "upstox_background.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(_log_file, encoding="utf-8")
    ]
)
log = logging.getLogger(__name__)

# Colour palette (hex strings)
C_HDR_BG      = "1E1E2E"   # column-header row background
C_HDR_FG      = "CDD6F4"   # column-header row text
C_META_BG     = "EAF2FF"   # Fetch Time / Spot / Expiry / Strike / PCR
C_ATM_BG      = "FFF9C4"   # ATM row
C_ALT_BG      = "F8F9FA"   # OTM zebra (even rows)
C_ITM_CALL_BG = "A9DFBF"   # ITM call (strike < spot)
C_ITM_PUT_BG  = "F1948A"   # ITM put  (strike > spot)
C_OTM_CALL_BG = "D4EFDF"   # OTM call odd rows
C_OTM_PUT_BG  = "FADBD8"   # OTM put  odd rows
C_CALLS_TITLE = "1A5276"   # Calls sheet title bar
C_PUTS_TITLE  = "641E16"   # Puts  sheet title bar

# =============================================================================
#  COLUMN DEFINITIONS
# =============================================================================

_META_COLS = [
    ("Fetch Time",  "_fetch_time",  19, "YYYY-MM-DD HH:MM:SS"),
    ("Record Time", "_record_time", 19, "YYYY-MM-DD HH:MM:SS"),
    ("Expiry",      "_expiry",      12, "@"),
]

_CALL_COLS = [
    ("Call Volume",  "call_options.market_data.volume",     12, "#,##0"),
    ("Call OI",      "call_options.market_data.oi",         10, "#,##0"),
    ("Call Prev OI", "call_options.market_data.prev_oi",    12, "#,##0"),
    ("Call Chg OI",  "_call_chg_oi",                       12, "#,##0"),
    ("Call LTP",     "call_options.market_data.ltp",        10, "#,##0.00"),
    ("Call IV",      "call_options.option_greeks.iv",        9, "0.00"),
    ("Call Delta",   "call_options.option_greeks.delta",     9, "0.0000"),
    ("Call Gamma",   "call_options.option_greeks.gamma",    11, "0.00000"),
    ("Call Theta",   "call_options.option_greeks.theta",     9, "0.00"),
    ("Call Vega",    "call_options.option_greeks.vega",      9, "0.0000"),
    ("Call PoP %",   "call_options.option_greeks.pop",       9, "0.00"),
]

_STRIKE_COL = [
    ("Strike",       "strike_price",                       11, "#,##0.00"),
]

_PUT_COLS = [
    ("Put Volume",  "put_options.market_data.volume",     12, "#,##0"),
    ("Put OI",      "put_options.market_data.oi",         10, "#,##0"),
    ("Put Prev OI", "put_options.market_data.prev_oi",    12, "#,##0"),
    ("Put Chg OI",  "_put_chg_oi",                       12, "#,##0"),
    ("Put LTP",     "put_options.market_data.ltp",        10, "#,##0.00"),
    ("Put IV",      "put_options.option_greeks.iv",        9, "0.00"),
    ("Put Delta",   "put_options.option_greeks.delta",     9, "0.0000"),
    ("Put Gamma",   "put_options.option_greeks.gamma",    11, "0.00000"),
    ("Put Theta",   "put_options.option_greeks.theta",     9, "0.00"),
    ("Put Vega",    "put_options.option_greeks.vega",      9, "0.0000"),
    ("Put PoP %",   "put_options.option_greeks.pop",       9, "0.00"),
]

OPTION_CHAIN_COLS = _META_COLS + _CALL_COLS + _STRIKE_COL + _PUT_COLS

def update_env_token(new_token: str):
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    lines = []
    updated = False
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip().startswith("UPSTOX_ACCESS_TOKEN="):
                    lines.append(f"UPSTOX_ACCESS_TOKEN={new_token}\n")
                    updated = True
                else:
                    lines.append(line)
    if not updated:
        lines.append(f"UPSTOX_ACCESS_TOKEN={new_token}\n")
    with open(env_path, "w", encoding="utf-8") as f:
        f.writelines(lines)
    # Also update current session environment variable
    os.environ["UPSTOX_ACCESS_TOKEN"] = new_token
    log.info("New access token saved to .env and loaded into environment.")


def refresh_access_token_flow() -> bool:
    # Formulate authorization URL
    auth_url = (
        f"https://api.upstox.com/v2/login/authorization/dialog?"
        f"response_type=code&client_id={CLIENT_ID}&redirect_uri={REDIRECT_URI}"
    )
    log.info("Opening browser for Upstox authorization: %s", auth_url)
    
    # Try opening the web browser
    try:
        webbrowser.open(auth_url)
    except Exception as e:
        log.error("Failed to automatically open browser: %s", e)
        
    # Open Tkinter Dialogue Box to retrieve code
    try:
        # Create a hidden root Tk window to host dialogue
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        
        prompt_msg = (
            "Your Upstox session has expired or is invalid.\n\n"
            "A browser window has been opened for login.\n"
            "If it did not open automatically, please visit this URL:\n"
            f"{auth_url}\n\n"
            "After logging in, copy the 'code' parameter from the address bar\n"
            "(e.g., from the URL: https://www.google.com/?code=XXXX)\n"
            "and paste it below:"
        )
        
        code = simpledialog.askstring(
            title="Upstox Authentication Required",
            prompt=prompt_msg,
            parent=root
        )
        root.destroy()
        
        if not code:
            log.warning("Authentication cancelled by user (no code entered).")
            return False
            
        code = code.strip()
        log.info("Received authorization code: %s. Exchanging for access token...", code)
        
        # Send post request to exchange code for token
        token_url = "https://api.upstox.com/v2/login/authorization/token"
        headers = {
            "accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded"
        }
        data = {
            "code": code,
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "redirect_uri": REDIRECT_URI,
            "grant_type": "authorization_code"
        }
        
        r = requests.post(token_url, headers=headers, data=data, timeout=15)
        r.raise_for_status()
        res_data = r.json()
        
        new_token = res_data.get("access_token")
        if not new_token:
            log.error("Exchange response did not contain access_token: %s", res_data)
            return False
            
        # Save token to .env
        update_env_token(new_token)
        return True
        
    except Exception as e:
        log.error("Error during authentication flow: %s", e)
        try:
            # Show friendly error dialog
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Upstox Authentication Error", f"Authentication failed:\n{e}")
            root.destroy()
        except Exception:
            pass
        return False


# =============================================================================
#  API LAYER
# =============================================================================

def _headers():
    return {"Authorization": f"Bearer {get_access_token()}", "Accept": "application/json"}


def fetch_expiries(instrument_key):
    global token_invalid
    url = f"{BASE_URL}/option/contract"
    try:
        r = requests.get(url, headers=_headers(),
                         params={"instrument_key": instrument_key}, timeout=15)
        if r.status_code == 401:
            token_invalid = True
        r.raise_for_status()
        expiries = sorted(set(c["expiry"] for c in r.json().get("data", []) if c.get("expiry")))
        log.info("Found %d expiry dates for %s", len(expiries), instrument_key)
        return expiries
    except Exception as e:
        if isinstance(e, requests.HTTPError) and e.response is not None and e.response.status_code == 401:
            token_invalid = True
        log.error("fetch_expiries error: %s", e)
        return []


def fetch_option_chain(instrument_key, expiry_date):
    url = f"{BASE_URL}/option/chain"
    r = requests.get(url, headers=_headers(),
                     params={"instrument_key": instrument_key, "expiry_date": expiry_date},
                     timeout=15)
    r.raise_for_status()
    body = r.json()
    if body.get("status") != "success":
        raise ValueError(f"API non-success: {body}")
    return body.get("data", []), r.headers.get("Date")


def get_nearest_expiry(instrument_key):
    expiries = fetch_expiries(instrument_key)
    today    = datetime.date.today().isoformat()
    upcoming = [e for e in expiries if e >= today]
    if not upcoming:
        raise RuntimeError("No upcoming expiries found.")
    chosen = upcoming[0]
    log.info("Auto-selected expiry: %s  (available: %s)", chosen, expiries[:5])
    return chosen

# =============================================================================
#  DATA HELPERS
# =============================================================================

def deep_get(obj, path):
    for part in path.split("."):
        if not isinstance(obj, dict):
            return None
        obj = obj.get(part)
    return obj


def parse_server_date(date_str):
    if not date_str:
        return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        dt = email.utils.parsedate_to_datetime(date_str)
        local_dt = dt.astimezone()
        return local_dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def enrich_row(row):
    row["_call_chg_oi"] = (deep_get(row, "call_options.market_data.oi")      or 0) - \
                          (deep_get(row, "call_options.market_data.prev_oi") or 0)
    row["_put_chg_oi"]  = (deep_get(row, "put_options.market_data.oi")       or 0) - \
                          (deep_get(row, "put_options.market_data.prev_oi")  or 0)
    return row


def filter_strikes(data, spot, n):
    if n is None or spot is None or not data:
        return data
    strikes = [r["strike_price"] for r in data]
    atm     = min(range(len(strikes)), key=lambda i: abs(strikes[i] - spot))
    return data[max(0, atm - n) : min(len(data) - 1, atm + n) + 1]


def get_spot_price(data):
    return data[0].get("underlying_spot_price") if data else None


def _atm_strike(data, spot):
    if spot and data:
        strikes = [r["strike_price"] for r in data]
        return min(strikes, key=lambda s: abs(s - spot))
    return None

# =============================================================================
#  COLOUR HELPERS
# =============================================================================

def _hex_to_rgb(h):
    """Convert '#RRGGBB' or 'RRGGBB' to (R, G, B) tuple."""
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

# =============================================================================
#  PATH A — openpyxl  (file not open in Excel)
# =============================================================================

def _op_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _op_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


def _op_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _op_write_header(ws, cols, title, title_bg):
    """Write two header rows via openpyxl. Called once per sheet."""
    thin  = _op_border()
    total = len(cols)

    # Row 1 — merged title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, color=C_HDR_FG, size=12, name="Calibri")
    c.fill      = _op_fill(title_bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — column headers
    for ci, (label, _, width, _) in enumerate(cols, start=1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font      = _op_font(bold=True, color=C_HDR_FG)
        c.fill      = _op_fill(C_HDR_BG)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"


def _op_append_rows(ws, data, spot, expiry, fetched_at, server_date_str, cols, overwrite=False):
    """Append or overwrite new rows via openpyxl."""
    thin  = _op_border()
    atm   = _atm_strike(data, spot)

    num_data_rows = len(data)
    expected_total_rows = 2 + num_data_rows
    record_time = parse_server_date(server_date_str)

    if overwrite and ws.max_row == expected_total_rows:
        log.info("openpyxl: Sheet '%s' already styled. Updating values only.", ws.title)
        for off, raw in enumerate(data):
            row = enrich_row(raw)
            row["_fetch_time"] = fetched_at
            row["_record_time"] = record_time
            row["_spot"]       = spot
            row["_expiry"]     = expiry
            row_idx = 3 + off
            for ci, (_, key, _, _) in enumerate(cols, start=1):
                c = ws.cell(row=row_idx, column=ci)
                c.value = row.get(key) if key.startswith("_") else deep_get(row, key)
        return len(data)

    if overwrite and ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    start = ws.max_row + 1

    # Dynamic indices for column categories
    meta_start, meta_end = 1, len(_META_COLS)
    call_start, call_end = meta_end + 1, meta_end + len(_CALL_COLS)
    strike_start, strike_end = call_end + 1, call_end + len(_STRIKE_COL)
    put_start, put_end = strike_end + 1, strike_end + len(_PUT_COLS)

    for off, raw in enumerate(data):
        row               = enrich_row(raw)
        row["_fetch_time"] = fetched_at
        row["_record_time"] = record_time
        row["_spot"]       = spot
        row["_expiry"]     = expiry
        strike      = row.get("strike_price", 0)
        is_atm      = atm is not None and strike == atm
        row_idx     = start + off
        alt         = row_idx % 2 == 0

        for ci, (_, key, _, fmt) in enumerate(cols, start=1):
            c = ws.cell(row=row_idx, column=ci)
            c.value        = row.get(key) if key.startswith("_") else deep_get(row, key)
            c.number_format = fmt
            c.alignment    = Alignment(horizontal="center", vertical="center")
            c.border       = thin
            c.font         = _op_font(bold=is_atm, size=9)

            if is_atm:
                c.fill = _op_fill(C_ATM_BG)
            else:
                if meta_start <= ci <= meta_end or strike_start <= ci <= strike_end:
                    c.fill = _op_fill(C_META_BG)
                elif call_start <= ci <= call_end:
                    if spot is not None and strike < spot:
                        c.fill = _op_fill(C_ITM_CALL_BG)
                    else:
                        c.fill = _op_fill(C_ALT_BG if alt else C_OTM_CALL_BG)
                elif put_start <= ci <= put_end:
                    if spot is not None and strike > spot:
                        c.fill = _op_fill(C_ITM_PUT_BG)
                    else:
                        c.fill = _op_fill(C_ALT_BG if alt else C_OTM_PUT_BG)

        ws.row_dimensions[row_idx].height = 15

    return len(data)


def _op_ensure_greeks(wb):
    if "Greeks Guide" in wb.sheetnames:
        return
    ws = wb.create_sheet("Greeks Guide")
    rows = [
        ("Greek",    "Meaning",                                                        "Typical Range"),
        ("Delta",    "Change in option price per ₹1 move in underlying",               "0→1 (call) | -1→0 (put)"),
        ("Gamma",    "Rate of change of Delta per ₹1 move",                           "Always +ve; peaks at ATM"),
        ("Theta",    "Daily time decay — option price lost per day",                   "Always -ve (₹/day)"),
        ("Vega",     "Price change per 1% change in Implied Volatility",               "Always +ve"),
        ("IV",       "Implied Volatility — market's future vol expectation",           "% annualised"),
        ("PoP %",    "Probability of Profit at expiry",                                "0–100 % (≈50 at ATM)"),
        ("PCR",      "Put-Call OI Ratio — >1 bearish, <1 bullish",                    ">1 / <1"),
        ("Chg OI",   "Change in Open Interest vs. previous close",                     "+ve = new longs added"),
        ("ITM Call", "Strike < Spot  →  call is in-the-money  (green)",               "—"),
        ("ITM Put",  "Strike > Spot  →  put  is in-the-money  (red)",                 "—"),
        ("ATM",      "Strike nearest to Spot  (yellow row)",                           "—"),
    ]
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    for ri, rd in enumerate(rows, start=1):
        for ci, val in enumerate(rd, start=1):
            cell            = ws.cell(row=ri, column=ci, value=val)
            cell.font       = _op_font(bold=(ri == 1),
                                       color=C_HDR_FG if ri == 1 else "000000", size=10)
            if ri == 1:
                cell.fill   = _op_fill(C_HDR_BG)
            cell.alignment  = Alignment(wrap_text=True, vertical="center")
            cell.border     = _op_border()
        ws.row_dimensions[ri].height = 20


def _openpyxl_run_multi(expiry_data_list, is_history=False):
    """Full openpyxl path: load → write all expiries → save once."""
    filepath = HISTORY_FILE if is_history else OUTPUT_FILE
    
    if os.path.exists(filepath):
        try:
            wb = openpyxl.load_workbook(filepath)
            is_new = False
        except Exception as e:
            log.warning("Could not open existing file (%s) — creating fresh.", e)
            wb = openpyxl.Workbook()
            is_new = True
    else:
        wb = openpyxl.Workbook()
        is_new = True

    for expiry, data, spot, fetched_at, server_date_str in expiry_data_list:
        sheet_name = expiry
        oc_new = False
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name, 0)
            oc_new = True
        
        oc_ws = wb[sheet_name]
        
        title_oc = f"OPTION CHAIN  |  {UNDERLYING}  |  Expiry: {expiry}"
        if is_history:
            title_oc += "  (HISTORY)"

        if oc_new or oc_ws.max_row < 2:
            _op_write_header(oc_ws, OPTION_CHAIN_COLS, title_oc, C_CALLS_TITLE)

        _op_append_rows(oc_ws, data, spot, expiry, fetched_at, server_date_str, OPTION_CHAIN_COLS, overwrite=not is_history)
        log.info("openpyxl: Prepared sheet '%s' (%d rows)", sheet_name, oc_ws.max_row)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    _op_ensure_greeks(wb)

    wb.save(filepath)
    wb.close()
    log.info("Saved openpyxl -> %s", filepath)
    return True

# =============================================================================
#  PATH B — xlwings  (file is open in Excel — write via COM, NO file touching)
# =============================================================================

def _xw_last_row(xw_sheet):
    """Return the last used row index (1-based). Returns 0 for empty sheet."""
    try:
        used = xw_sheet.api.UsedRange
        return used.Row + used.Rows.Count - 1
    except Exception:
        return 0


def _xw_write_header(xw_sheet, cols, title, title_bg_hex):
    """Write two header rows directly via xlwings COM."""
    n = len(cols)
    xl_center = -4108  # xlCenter

    # Row 1 — merged title
    try:
        xw_sheet.range((1, 1), (1, n)).api.Merge()
    except Exception:
        pass
    xw_sheet.range(1, 1).value = title
    xw_sheet.range((1, 1), (1, n)).color     = _hex_to_rgb(title_bg_hex)
    xw_sheet.range(1, 1).font.bold           = True
    xw_sheet.range(1, 1).font.size           = 12
    xw_sheet.range(1, 1).font.color         = _hex_to_rgb(C_HDR_FG)
    xw_sheet.range(1, 1).api.HorizontalAlignment = xl_center
    xw_sheet.api.Rows(1).RowHeight = 22

    # Row 2 — column labels
    xw_sheet.range(2, 1).value = [[col[0] for col in cols]]
    hdr = xw_sheet.range((2, 1), (2, n))
    hdr.color                     = _hex_to_rgb(C_HDR_BG)
    hdr.font.bold                 = True
    hdr.font.color                = _hex_to_rgb(C_HDR_FG)
    hdr.api.HorizontalAlignment   = xl_center
    hdr.api.WrapText              = True
    xw_sheet.api.Rows(2).RowHeight = 28

    # Column widths
    for ci, (_, _, width, _) in enumerate(cols, start=1):
        xw_sheet.api.Columns(ci).ColumnWidth = width

    # Freeze at row 3
    try:
        xw_sheet.activate()
        win = xw_sheet.book.app.api.ActiveWindow
        win.FreezePanes = False
        xw_sheet.range("A3").api.Select()
        win.FreezePanes = True
    except Exception:
        pass


def _xw_ensure_headers(xw_sheet, cols, title, title_bg_hex):
    """Write header rows only if the sheet is empty or missing them."""
    try:
        if xw_sheet.range(2, 1).value == cols[0][0]:
            return   # headers already present
    except Exception:
        pass
    _xw_write_header(xw_sheet, cols, title, title_bg_hex)


def _xw_append_rows(xw_sheet, data, spot, expiry, fetched_at, server_date_str, cols, overwrite=False):
    """
    Append or overwrite rows directly to an open Excel sheet via xlwings COM.
    Uses bulk value write (1 COM call) + range-based colour application.
    """
    xl_center = -4108   # xlCenter constant

    last_row  = _xw_last_row(xw_sheet)
    num_data_rows = len(data)
    expected_total_rows = 2 + num_data_rows

    record_time = parse_server_date(server_date_str)

    # ── Enrich all rows and build value matrix ────────────────────────────────
    enriched = []
    matrix   = []
    for raw in data:
        row               = enrich_row(raw)
        row["_fetch_time"] = fetched_at
        row["_record_time"] = record_time
        row["_spot"]       = spot
        row["_expiry"]     = expiry
        enriched.append(row)
        matrix.append([
            row.get(key) if key.startswith("_") else deep_get(row, key)
            for (_, key, _, _) in cols
        ])

    if overwrite and last_row == expected_total_rows:
        # Just update values in one bulk COM call!
        log.info("xlwings: Sheet '%s' already styled. Updating values only.", xw_sheet.name)
        xw_sheet.range(3, 1).value = matrix
        return len(data)

    if overwrite and last_row >= 3:
        try:
            xw_sheet.range((3, 1), (last_row, len(cols))).clear()
        except Exception as e:
            log.warning("xlwings clear error: %s", e)
        last_row = 2

    start_row = last_row + 1
    n_cols    = len(cols)
    atm       = _atm_strike(data, spot)

    # Dynamic indices for column categories
    meta_start, meta_end = 1, len(_META_COLS)
    call_start, call_end = meta_end + 1, meta_end + len(_CALL_COLS)
    strike_start, strike_end = call_end + 1, call_end + len(_STRIKE_COL)
    put_start, put_end = strike_end + 1, strike_end + len(_PUT_COLS)

    # ── 1 COM call: write all values at once ──────────────────────────────────
    xw_sheet.range(start_row, 1).value = matrix

    # ── Number formats (one call per column) ──────────────────────────────────
    end_row = start_row + len(data) - 1
    for ci, (_, _, _, fmt) in enumerate(cols, start=1):
        xw_sheet.range(start_row, ci).resize(len(data), 1).number_format = fmt
    xw_sheet.range((start_row, 1), (end_row, n_cols)) \
            .api.HorizontalAlignment = xl_center

    # ── Borders for the entire new block (1 COM call) ─────────────────────────
    block = xw_sheet.range((start_row, 1), (end_row, n_cols)).api
    block.Borders.LineStyle = 1   # xlContinuous
    block.Borders.Weight    = 2   # xlThin

    # ── Row height (1 COM call) ───────────────────────────────────────────────
    block.RowHeight = 15

    # ── Apply colours by group (one range call per row) ───────────────────────
    for off, row in enumerate(enriched):
        strike  = row.get("strike_price", 0)
        ri = start_row + off
        is_atm  = atm is not None and strike == atm
        alt     = ri % 2 == 0

        if is_atm:
            # Entire row ATM color
            xw_sheet.range((ri, 1), (ri, n_cols)).color = _hex_to_rgb(C_ATM_BG)
            xw_sheet.range((ri, 1), (ri, n_cols)).font.bold = True
        else:
            # Meta columns (Fetch Time, Expiry)
            xw_sheet.range((ri, meta_start), (ri, meta_end)).color = _hex_to_rgb(C_META_BG)
            # Strike column
            xw_sheet.range((ri, strike_start), (ri, strike_end)).color = _hex_to_rgb(C_META_BG)

            # Call columns
            if spot is not None and strike < spot:
                xw_sheet.range((ri, call_start), (ri, call_end)).color = _hex_to_rgb(C_ITM_CALL_BG)
            else:
                c_bg = C_ALT_BG if alt else C_OTM_CALL_BG
                xw_sheet.range((ri, call_start), (ri, call_end)).color = _hex_to_rgb(c_bg)

            # Put columns
            if spot is not None and strike > spot:
                xw_sheet.range((ri, put_start), (ri, put_end)).color = _hex_to_rgb(C_ITM_PUT_BG)
            else:
                p_bg = C_ALT_BG if alt else C_OTM_PUT_BG
                xw_sheet.range((ri, put_start), (ri, put_end)).color = _hex_to_rgb(p_bg)

    return len(data)


def _xw_ensure_greeks(book):
    """Add Greeks Guide sheet directly via xlwings if missing."""
    if "Greeks Guide" in [s.name for s in book.sheets]:
        return
    try:
        ws = book.sheets.add("Greeks Guide")
        rows = [
            ("Greek",    "Meaning",                                                  "Typical Range"),
            ("Delta",    "Change in option price per ₹1 move in underlying",         "0→1 (call) | -1→0 (put)"),
            ("Gamma",    "Rate of change of Delta per ₹1 move",                     "Always +ve"),
            ("Theta",    "Daily time decay",                                         "Always -ve (₹/day)"),
            ("Vega",     "Price change per 1% IV change",                            "Always +ve"),
            ("IV",       "Implied Volatility",                                       "% annualised"),
            ("PoP %",    "Probability of Profit at expiry",                          "0–100 %"),
            ("PCR",      "Put-Call OI Ratio",                                        ">1 bearish | <1 bullish"),
            ("Chg OI",   "Change in OI vs. previous close",                          "+ve = new longs"),
            ("ITM Call", "Strike < Spot (green)",                                    "—"),
            ("ITM Put",  "Strike > Spot (red)",                                      "—"),
            ("ATM",      "Strike nearest Spot (yellow)",                             "—"),
        ]
        ws.range(1, 1).value = rows
        ws.range((1, 1), (1, 3)).color     = _hex_to_rgb(C_HDR_BG)
        ws.range((1, 1), (1, 3)).font.bold = True
        ws.range((1, 1), (1, 3)).font.color = _hex_to_rgb(C_HDR_FG)
        ws.api.Columns(1).ColumnWidth = 14
        ws.api.Columns(2).ColumnWidth = 60
        ws.api.Columns(3).ColumnWidth = 30
    except Exception as e:
        log.debug("Greeks Guide via xlwings skipped: %s", e)


def _xlwings_run_multi(expiry_data_list, is_history=False):
    """Write directly to the open Excel workbook via xlwings COM. No file I/O."""
    try:
        import xlwings as xw  # type: ignore
    except ImportError:
        log.error("xlwings not installed — run: pip install xlwings")
        return False

    filepath = HISTORY_FILE if is_history else OUTPUT_FILE
    filename = os.path.basename(filepath)
    book = None
    try:
        for b in xw.books:
            if b.name.lower() == filename.lower() or \
               b.fullname.lower() == filepath.lower():
                book = b
                break
    except Exception:
        pass

    if book is None:
        try:
            book = xw.Book(filepath)
        except Exception as e:
            log.error("Cannot open via xlwings: %s", e)
            return False

    for expiry, data, spot, fetched_at, server_date_str in expiry_data_list:
        title_oc = f"OPTION CHAIN  |  {UNDERLYING}  |  Expiry: {expiry}"
        if is_history:
            title_oc += "  (HISTORY)"

        sheet_name = expiry
        try:
            xws = book.sheets[sheet_name]
        except Exception:
            xws = book.sheets.add(sheet_name)

        _xw_ensure_headers(xws, OPTION_CHAIN_COLS, title_oc, C_CALLS_TITLE)
        n = _xw_append_rows(xws, data, spot, expiry, fetched_at, server_date_str, OPTION_CHAIN_COLS, overwrite=not is_history)
        log.info("xlwings: Prepared sheet '%s' (%s, total row %d)", sheet_name, filename, _xw_last_row(xws))

    _xw_ensure_greeks(book)
    book.save()
    log.info("Saved via xlwings -> %s", filepath)
    return True


# =============================================================================
#  FILE LOCK CHECK
# =============================================================================

def _is_locked(filepath):
    """True if the file exists and is currently locked (e.g., open in Excel)."""
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, "r+b"):
            return False
    except (PermissionError, IOError):
        return True


# =============================================================================
#  MAIN LOOP
# =============================================================================

def fetch_expiry_chain(expiry):
    global token_invalid
    fetched_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log.info("Fetching | %s | expiry=%s", UNDERLYING, expiry)
    try:
        data, server_date_str = fetch_option_chain(UNDERLYING, expiry)
        if not data:
            log.warning("Empty response for expiry %s — market may be closed.", expiry)
            return None
        spot = get_spot_price(data)
        log.info("Spot: %s  |  Strikes fetched: %d for expiry %s", spot, len(data), expiry)
        data = filter_strikes(data, spot, STRIKES_AROUND_ATM)
        return (expiry, data, spot, fetched_at, server_date_str)
    except Exception as e:
        if isinstance(e, requests.HTTPError) and e.response is not None and e.response.status_code == 401:
            token_invalid = True
        log.error("Fetch error for expiry %s: %s", expiry, e)
        return None


def main():
    global token_invalid

    log.info("=" * 65)
    log.info("  Upstox Live Option Chain (Multi-Expiry)")
    log.info("  Underlying : %s", UNDERLYING)
    log.info("  Refresh    : every %d s", REFRESH_INTERVAL)
    log.info("  Output     : %s", OUTPUT_FILE)
    log.info("  History    : %s", HISTORY_FILE)
    log.info("  Strikes    : %s around ATM", STRIKES_AROUND_ATM or "ALL")
    log.info("=" * 65)

    # Initialize check on startup
    token = get_access_token()
    if token == "YOUR_ACCESS_TOKEN_HERE" or not token:
        token_invalid = True

    while True:
        if token_invalid:
            log.warning("Access token is invalid, expired, or missing. Triggering login flow...")
            success = refresh_access_token_flow()
            if success:
                token_invalid = False
            else:
                log.warning("Login flow did not complete successfully. Retrying in 60s...")
                time.sleep(60)
                continue
        # Determine expiries to fetch
        if EXPIRY_DATE:
            expiries_to_fetch = [EXPIRY_DATE]
        else:
            try:
                all_expiries = fetch_expiries(UNDERLYING)
                today = datetime.date.today().isoformat()
                expiries_to_fetch = [e for e in all_expiries if e >= today]
            except Exception as e:
                log.error("Could not fetch expiry dates: %s", e)
                expiries_to_fetch = []

        if not expiries_to_fetch:
            log.warning("No expiry dates to fetch in this cycle.")
        else:
            log.info("Starting concurrent refresh cycle for %d expiry dates...", len(expiries_to_fetch))
            
            results = []
            with ThreadPoolExecutor(max_workers=min(10, len(expiries_to_fetch))) as executor:
                future_to_expiry = {executor.submit(fetch_expiry_chain, expiry): expiry for expiry in expiries_to_fetch}
                for future in as_completed(future_to_expiry):
                    expiry = future_to_expiry[future]
                    try:
                        res = future.result()
                        if res:
                            results.append(res)
                    except Exception as e:
                        log.error("Exception during fetch for expiry %s: %s", expiry, e)
            
            if results:
                # Sort results by expiry date (ascending)
                results.sort(key=lambda x: x[0])
                
                # 1. Update/overwrite the live file
                try:
                    if _is_locked(OUTPUT_FILE):
                        _xlwings_run_multi(results, is_history=False)
                    else:
                        _openpyxl_run_multi(results, is_history=False)
                except Exception as e:
                    log.error("Failed writing live file: %s", e)

                # 2. Append to the history file
                try:
                    if _is_locked(HISTORY_FILE):
                        _xlwings_run_multi(results, is_history=True)
                    else:
                        _openpyxl_run_multi(results, is_history=True)
                except Exception as e:
                    log.error("Failed writing history file: %s", e)

        log.info("Next refresh cycle in %d s …", REFRESH_INTERVAL)
        time.sleep(REFRESH_INTERVAL)


if __name__ == "__main__":
    main()
